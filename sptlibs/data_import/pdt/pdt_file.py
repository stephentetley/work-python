"""
Copyright 2024 Stephen Tetley

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

"""

import os
import glob
import openpyxl
import polars as pl
import duckdb
from jinja2 import Template
from openpyxl import load_workbook
from sptlibs.utils.xlsx_source import XlsxSource
import sptlibs.data_import.import_utils as import_utils


class PdtFile:
    def __init__(self, *, file_name: str, pdt_sheet_name: str, sys_floc: str, sys_name: str):
        self.file_name = file_name
        self.pdt_type = pdt_sheet_name
        self.system_floc = sys_floc
        self.system_name = sys_name

    def __str__(self) -> str:
        return f"<<file_name='{self.file_name}', system_floc='{self.system_floc}', pdt_type='{self.pdt_type}'>>"
    
    @classmethod
    def from_file(cls, *, file_name: str, ):
        if os.path.exists(file_name):
            wb = load_workbook(filename=file_name)
            lead_sheet = wb['Asset Data Lead Sheet']            
            floc = _get_floc(lead_sheet=lead_sheet)
            sys_name = lead_sheet['B7'].value
            pdt_sheet = next(x for x in wb.sheetnames if x.startswith("PDT_"))
            wb.close()
            return cls(file_name=file_name, pdt_sheet_name=pdt_sheet, sys_floc=floc, sys_name=sys_name)
        else:
            return None

    def to_dict(self):
        return {"base_name": os.path.basename(self.file_name),
                "pdt_type": self.pdt_type,
                "system_floc": self.system_floc,
                "system_name": self.system_name,
                }
    
    def store_eav(self, *, con=duckdb.DuckDBPyConnection) -> None:
        pdt_src = XlsxSource(self.file_name, self.pdt_type)
        eav_df = import_utils.read_xlsx_source(pdt_src, normalize_column_names=True)
        con.register(view_name="eav_df", python_object=eav_df)
        con.execute(Template(_unpivot_template).render(base_name=os.path.basename(self.file_name))).pl()

_unpivot_template = """
    INSERT INTO pdt_raw_data.pdt_eav BY NAME
    WITH cte AS (
        UNPIVOT eav_df ON COLUMNS (* EXCLUDE asset_name) INTO NAME entity_name VALUE attr_value
    )
    SELECT '{{base_name}}' AS base_name, t.entity_name AS entity_name, t.asset_name AS attr_name, t.attr_value AS attr_value FROM cte t ORDER BY t.entity_name, t.asset_name;
"""


def _get_floc(lead_sheet: openpyxl.worksheet.worksheet.Worksheet) -> str:
    floc_parts = [
        lead_sheet['C3'].value, 
        lead_sheet['C4'].value, 
        lead_sheet['C5'].value,
        lead_sheet['C6'].value,
        lead_sheet['C7'].value,
    ]
    floc_trim = []
    for x in floc_parts:
        if x:
            floc_trim.append(x)
        else: 
            break

    floc = "-".join(floc_trim)
    return floc


class PdtFiles:
    def __init__(self, *, pdt_files: list[PdtFile]):
        self.pdt_files = pdt_files
    
    def __iter__(self):
        return iter(self.pdt_files)

    def __len__(self):
         return len(self.pdt_files)

    @classmethod
    def from_files(cls, *, src_dir: str):
        if os.path.exists(src_dir):
            globlist = glob.glob("*.xlsx", root_dir=src_dir, recursive=False)
            def full_path(file_name): 
                return os.path.normpath(os.path.join(src_dir, file_name))
            pdt_files = [PdtFile.from_file(file_name=full_path(e)) for e in globlist]
            return cls(pdt_files=pdt_files)
        else:
            return None
        
    def to_df(self) -> pl.DataFrame:
        schema = {"base_name": pl.String, "pdt_type": pl.String,
                    "system_floc": pl.String, "system_name": pl.String}
        rows = [e.to_dict() for e in self.pdt_files]
        return pl.from_dicts(rows, schema=schema)
    
    def store(self, *, con=duckdb.DuckDBPyConnection) -> None:
        summary_df = self.to_df()
        con.register(view_name="summary_df", python_object=summary_df)
        insert_stmt = "INSERT INTO pdt_raw_data.pdt_file SELECT * FROM summary_df"
        con.execute(insert_stmt)
        for e in self.pdt_files:
            e.store_eav(con=con)

        