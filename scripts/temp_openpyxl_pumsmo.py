from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from openpyxl.utils import quote_sheetname, absolute_coordinate

# Write directly with openpyxl
# Trying to write from Pandas gives away control over validation

# Ideally column_index and ws_meta would be encapsulated state
def add_enums_table(*, 
                    table_name: str,
                    column_header: str,
                    column_index: int,
                    enum_values: list[str],
                    ws_meta: Worksheet) -> None:
    ws_meta.cell(row=1, column=column_index).value = column_header
    for ix, s in enumerate(enum_values):
        ws_meta.cell(row=ix+2, column=column_index).value = s
    column_letter = get_column_letter(column_index)
    ref_string = f"{column_letter}1:{column_letter}{ix+2}"
    print(f"{column_header} {ref_string}")
    tab = Table(displayName=table_name, ref=ref_string)
    ws_meta.add_table(tab)

wb = Workbook()
ws = wb.active
ws.title = "pumsmo"
wb.create_sheet("metadata_tables")

md = wb["metadata_tables"]


add_enums_table(table_name='TABLE1',
                column_header='INSULATION_CLASS_DEG_C',
                column_index=1,
                enum_values= ['200 (CLASS K)',
                              '180 (CLASS H)',
                              '155 (CLASS F)',
                              '130 (CLASS B)',
                              '120 (CLASS E)',
                              '105 (CLASS A)'
                              ],
                ws_meta=md)

ref =  f"{quote_sheetname(md.title)}!{absolute_coordinate('A2:A7')}"
defn = DefinedName("insulation_class_range", attr_text=ref)
wb.defined_names.add(defn)

ws.cell(row=1, column=1).value = 'Insulation Class'
dv = DataValidation(type="list", formula1='insulation_class_range', allow_blank=True)
ws.add_data_validation(dv)
dv.add('A2:A1000')


add_enums_table(table_name='TABLE2',
                column_header='PUMS_LIFTING_TYPE',
                column_index=2,
                enum_values= ['BLUE ROPE',
                              'CHAIN',
                              'DAVIT ARM'
                              ],
                ws_meta=md)

ws.cell(row=1, column=2).value = 'Lifting Type'
dv = DataValidation(type="list", formula1='INDIRECT("TABLE2[PUMS_LIFTING_TYPE]")', allow_blank=True)
ws.add_data_validation(dv)
dv.add('B2:B1000')

# Use Defined names rather than tables, see
# https://openpyxl.readthedocs.io/en/stable/defined_names.html


add_enums_table(table_name='TABLE3',
                column_header='CATEGORY',
                column_index=3,
                enum_values= ['E',
                              'H',
                              'I'
                              ],
                ws_meta=md)

ref =  f"{quote_sheetname(md.title)}!{absolute_coordinate('C2:C5')}"
defn = DefinedName("category_range", attr_text=ref)
wb.defined_names.add(defn)

add_enums_table(table_name='TABLE4',
                column_header='E_OBJTYPE',
                column_index=4,
                enum_values= ['POGE',
                              'PODE',
                              'METER'
                              ],
                ws_meta=md)

ref =  f"{quote_sheetname(md.title)}!{absolute_coordinate('D2:D5')}"
defn = DefinedName("E_OBJTYPE", attr_text=ref)
wb.defined_names.add(defn)


add_enums_table(table_name='TABLE5',
                column_header='H_OBJTYPE',
                column_index=5,
                enum_values= ['LDSS'
                              ],
                ws_meta=md)

ref =  f"{quote_sheetname(md.title)}!{absolute_coordinate('E2:E3')}"
defn = DefinedName("H_OBJTYPE", attr_text=ref)
wb.defined_names.add(defn)


add_enums_table(table_name='TABLE6',
                column_header='I_OBJTYPE',
                column_index=6,
                enum_values= ['FSTN',
                              'LSTN'
                              ],
                ws_meta=md)

ref =  f"{quote_sheetname(md.title)}!{absolute_coordinate('F2:F4')}"
defn = DefinedName("I_OBJTYPE", attr_text=ref)
wb.defined_names.add(defn)


ws.cell(row=1, column=3).value = 'Category'
dv = DataValidation(type="list", formula1='category_range', allow_blank=True)
ws.add_data_validation(dv)
dv.add('C2:C1000')

ws.cell(row=1, column=4).value = 'ObjType'
dv = DataValidation(type="list", formula1='INDIRECT(C2 & "_OBJTYPE")', allow_blank=True)
ws.add_data_validation(dv)
dv.add('D2:D1000')

wb.save("pumsmo2.xlsx")

print(f"{quote_sheetname('metadata_tables')}")
print(f"{absolute_coordinate('A1:A5')}")


