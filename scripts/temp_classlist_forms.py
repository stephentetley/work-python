
import os
import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from openpyxl.utils import quote_sheetname, absolute_coordinate
from openpyxl.utils.dataframe import dataframe_to_rows

import sptlibs.data_access.s4_classlists.s4_classlists_import as s4_classlists_import
import sptapps.classlist_forms.make_classlist_forms as make_classlist_forms

# TODO classlists should include ordering information...
s4_classlists_source = 'g:/work/2025/asset_data_facts/s4_classlists/s4_classlists_jan2025.duckdb'
duckdb_path = 'e:/coding/work/work-sql/classlist_forms/output1.duckdb'
xlsx_path = 'e:/coding/work/work-sql/classlist_forms/lstnut.xlsx'

if os.path.exists(duckdb_path):
    os.remove(duckdb_path)

con = duckdb.connect(database=duckdb_path, read_only=False)

s4_classlists_import.copy_classlists_tables(classlists_source_db_path=s4_classlists_source, setup_tables=True, dest_con=con)

wb = Workbook()
ws = wb.active
ws.title = "lstnut"
wb.create_sheet("lstnut_metadata")

md = wb["lstnut_metadata"]

df = make_classlist_forms.get_class_enums(class_name='LSTNUT', con=con)
for df_row in dataframe_to_rows(df, index=False, header=True):
    md.append(df_row)

df = make_classlist_forms.get_class_enums_dimensions(class_name='LSTNUT', con=con)
for ix, df_row in df.iterrows():
    column_letter = get_column_letter(df_row['column_idx'])
    range = f"{column_letter}2:{column_letter}{df_row['enum_count']+1}"
    ref =  f"{quote_sheetname(md.title)}!{absolute_coordinate(range)}"
    defn = DefinedName(df_row['range_name'], attr_text=ref)
    wb.defined_names.add(defn)
    print(df_row['range_name'], range, ref)

df = make_classlist_forms.get_char_form_data(class_name='LSTNUT', con=con)
for ix, df_row in df.iterrows():
    col = df_row['column_idx']
    column_letter = get_column_letter(df_row['column_idx'])
    range = f"{column_letter}2:{column_letter}1000"
    ws.cell(row=1, column=col).value = df_row['column_heading']
    ws.column_dimensions[column_letter].width = df_row['column_width']
    dv = DataValidation(type=df_row['validation_type'], 
                        operator=df_row['validation_operator'], 
                        formula1=df_row['validation_formula1'],
                        allow_blank=True)
    dv.promptTitle = df_row['validation_prompt_title']
    dv.prompt = df_row['validation_prompt']
    dv.showInputMessage = True
    dv.errorTitle = df_row['validation_error_title']
    dv.error =df_row['validation_error']
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(range)
    print(range)

ws.freeze_panes = 'A2'




wb.save(xlsx_path)
wb.close()

con.close()

