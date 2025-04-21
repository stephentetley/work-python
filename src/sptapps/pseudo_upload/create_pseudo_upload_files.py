"""
Copyright 2025 Stephen Tetley

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

"""

import os
import duckdb
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from sptlibs.utils.sql_script_runner import SqlScriptRunner




def create_pseudo_upload_files(*, 
                               worklist_path: str,
                               loc_on_site_path: str,
                               output_root: str, 
                               con: duckdb.DuckDBPyConnection) -> None: 
    runner = SqlScriptRunner(__file__, con=con)
    runner.set_variable(var_name='worklist_file', var_value=worklist_path)
    runner.set_variable(var_name='loc_on_site_file', var_value=loc_on_site_path)
    runner.exec_sql_file(rel_file_path='setup_worklist_tables.sql')
    runner.exec_sql_file(rel_file_path='lstnut_unpivot_create_macros.sql')
    _output_files(output_root=output_root, con=con)


def set_bgcolour_of_range(*, range: str, colour: str, ws) -> None:
    for row in ws[range]:
        for cell in row:
            cell.fill = PatternFill(fill_type="solid", start_color=colour, end_color=colour)

def set_font_bold_of_range(*, range: str, ws) -> None:
    for row in ws[range]:
        for cell in row:
            cell.font = Font(bold=True)



    
def _output_files(*, output_root:str, con: duckdb.DuckDBPyConnection) -> None: 
    df = con.sql(query=_outer_select_stmt).pl()
    for row in df.rows(named=True):
        ai2_pli_num = row['ai2_pli_num']
        file_name = row['file_name']
        out_file = os.path.normpath(os.path.join(output_root, file_name))
        print(f"{out_file} ...")
        wb = Workbook()
        ws = wb.active
        ws.title = "master_data"
        # Master Data
        ws = wb["master_data"]
        ws.append(["Create new row in equipment list view..."])
        df = con.execute("SELECT * FROM get_masterdata($reference);", parameters={'reference': ai2_pli_num}).df()
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        ws.append([None])
        ws.append(["Enter the attr_values in the equipment details view..."])
        ws.append(["Remember, set [Data Origin] first"])
        df = con.execute("SELECT * FROM unpivot_masterdata($reference);", parameters={'reference': ai2_pli_num}).df()
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        ws.column_dimensions['A'].width = 55
        for ch in 'BCDEFG':
            ws.column_dimensions[ch].width = 30
        set_bgcolour_of_range(range='A2:G2', colour='00FFFF00', ws=ws)
        set_bgcolour_of_range(range='A7:B7', colour='00FFFF00', ws=ws)
        set_bgcolour_of_range(range='B8:B16', colour='00FFCC99', ws=ws)
        set_font_bold_of_range(range='A1:A1', ws=ws)
        set_font_bold_of_range(range='A5:A6', ws=ws)
        # LSTNUT
        wb.create_sheet("lstnut")
        ws = wb["lstnut"]
        ws.append(["Copy-paste the attr_values into class LSTNUT..."])
        ws.append(["The Characteristics table must be sorted by characteristic..."])
        df = con.execute("SELECT * FROM unpivot_lstnut($reference);", parameters={'reference': ai2_pli_num}).df()
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 40
        set_bgcolour_of_range(range='A3:B3', colour='00FFFF00', ws=ws)
        set_bgcolour_of_range(range='B4:B38', colour='00FFCC99', ws=ws)
        set_font_bold_of_range(range='A1:A2', ws=ws)
        # AIB_REFERENCE
        wb.create_sheet("aib_reference")
        ws = wb["aib_reference"]
        ws.append(["Copy-paste the attr_value into class AIB_REFERENCE..."])
        ws.append([f"Look up {ai2_pli_num} in AI2 to find its parent SAI number and add that."])
        query_aib = f"SELECT 'AI2_AIB_REFERENCE' AS attr_name, '{ai2_pli_num}' AS attr_value;"
        df = con.execute(query_aib).df()
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 40
        set_bgcolour_of_range(range='A3:B3', colour='00FFFF00', ws=ws)
        set_bgcolour_of_range(range='B4:B4', colour='00FFCC99', ws=ws)
        set_font_bold_of_range(range='A1:A2', ws=ws)
        # EAST_NORTH
        wb.create_sheet("east_north")
        ws = wb["east_north"]
        ws.append(["Copy-paste the attr_values into class EAST_NORTH..."])
        ws.append(["The Characteristics table must be sorted by characteristic..."])
        df = con.execute("SELECT * FROM unpivot_east_north($reference);", parameters={'reference': ai2_pli_num}).df()
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 40
        set_bgcolour_of_range(range='A3:B3', colour='00FFFF00', ws=ws)
        set_bgcolour_of_range(range='B4:B5', colour='00FFCC99', ws=ws)
        set_font_bold_of_range(range='A1:A2', ws=ws)
        # ASSET_CONDITION
        wb.create_sheet("asset_condition")
        ws = wb["asset_condition"]
        ws.append(["Copy-paste the attr_values into class ASSET_CONDITION..."])
        ws.append(["The Characteristics table must be sorted by characteristic..."])
        df = con.execute("SELECT * FROM unpivot_asset_condition($reference);", parameters={'reference': ai2_pli_num}).df()
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 40
        set_bgcolour_of_range(range='A3:B3', colour='00FFFF00', ws=ws)
        set_bgcolour_of_range(range='B4:B8', colour='00FFCC99', ws=ws)
        set_font_bold_of_range(range='A1:A2', ws=ws)
        wb.save(out_file)
        wb.close()

# TO remove WHERE clause when `_output_files` is working
_outer_select_stmt = """
SELECT 
    t.equipment_id AS ai2_pli_num,
    t.batch_name || '-' || t.s4_floc || '-' || t.equipment_id || '-' || udfx.make_snake_case_name(t.s4_name) || '.xlsx' AS file_name,
FROM worklist_extra.worklist t;
"""
